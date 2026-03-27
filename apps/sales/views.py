from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, F
from .models import Customer, SalesOrder, SalesOrderLine, Shipment


@login_required
def customer_list(request):
    q = request.GET.get('q', '')
    customers = Customer.objects.filter(is_active=True)
    if q:
        customers = customers.filter(Q(name__icontains=q) | Q(code__icontains=q))
    return render(request, 'sales/customer_list.html', {'customers': customers, 'q': q})


@login_required
def customer_detail(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    orders = customer.orders.all().order_by('-order_date')[:20]
    return render(request, 'sales/customer_detail.html', {'customer': customer, 'orders': orders})


@login_required
def customer_form(request, pk=None):
    customer = get_object_or_404(Customer, pk=pk) if pk else None
    if request.method == 'POST':
        data = request.POST
        if customer:
            for field in ['name', 'code', 'contact_name', 'email', 'phone', 'address', 'country']:
                setattr(customer, field, data.get(field, ''))
            customer.save()
            messages.success(request, "Client modifié.")
        else:
            customer = Customer.objects.create(
                name=data['name'], code=data['code'],
                contact_name=data.get('contact_name', ''),
                email=data.get('email', ''), phone=data.get('phone', ''),
                address=data.get('address', ''), country=data.get('country', ''),
            )
            messages.success(request, "Client créé.")
        return redirect('sales:customer_detail', pk=customer.pk)
    return render(request, 'sales/customer_form.html', {'customer': customer})


@login_required
def order_list(request):
    q = request.GET.get('q', '')
    status = request.GET.get('status', '')
    orders = SalesOrder.objects.select_related('customer').all()
    if q:
        orders = orders.filter(Q(order_number__icontains=q) | Q(customer__name__icontains=q))
    if status:
        orders = orders.filter(status=status)
    return render(request, 'sales/order_list.html', {
        'orders': orders, 'q': q, 'status': status,
        'status_choices': SalesOrder.STATUS_CHOICES
    })


@login_required
def order_detail(request, pk):
    order = get_object_or_404(SalesOrder, pk=pk)
    shipments = order.shipments.all()
    return render(request, 'sales/order_detail.html', {'order': order, 'shipments': shipments})


@login_required
def order_form(request, pk=None):
    order = get_object_or_404(SalesOrder, pk=pk) if pk else None
    customers = Customer.objects.filter(is_active=True)
    from apps.stock.models import Material
    finished_products = Material.objects.filter(material_type='finished_product', is_active=True)
    if request.method == 'POST':
        data = request.POST
        if order:
            order.customer_id = data['customer']
            order.status = data['status']
            order.order_date = data['order_date']
            order.delivery_date = data.get('delivery_date') or None
            order.discount = data.get('discount', 0)
            order.invoice_number = data.get('invoice_number', '')
            order.tva_rate = data.get('tva_rate', 19.0)
            order.payment_status = data.get('payment_status', 'unpaid')
            order.notes = data.get('notes', '')
            order.save()
        else:
            order = SalesOrder.objects.create(
                order_number=data['order_number'],
                customer_id=data['customer'],
                order_date=data['order_date'],
                delivery_date=data.get('delivery_date') or None,
                discount=data.get('discount', 0),
                invoice_number=data.get('invoice_number', ''),
                tva_rate=data.get('tva_rate', 19.0),
                payment_status=data.get('payment_status', 'unpaid'),
                notes=data.get('notes', ''),
                created_by=request.user
            )
        messages.success(request, "Commande enregistrée.")
        return redirect('sales:order_detail', pk=order.pk)
    return render(request, 'sales/order_form.html', {
        'order': order, 'customers': customers, 'finished_products': finished_products,
        'status_choices': SalesOrder.STATUS_CHOICES,
        'payment_choices': SalesOrder.PAYMENT_CHOICES
    })


@login_required
def order_invoice_pdf(request, pk):
    order = get_object_or_404(SalesOrder, pk=pk)
    return render(request, 'sales/invoice_print.html', {'order': order})


@login_required
def report_view(request):
    by_product = (
        SalesOrderLine.objects
        .values('material__name', 'material__code')
        .annotate(
            total_quantity=Sum('shipped_quantity'),
            total_revenue=Sum(F('shipped_quantity') * F('unit_price'))
        )
        .order_by('-total_revenue')
    )
    by_customer = (
        SalesOrder.objects
        .filter(status__in=['shipped', 'invoiced'])
        .values('customer__name', 'customer__code')
        .annotate(order_count=Sum('id'))
        .order_by('-order_count')
    )
    return render(request, 'sales/report.html', {
        'by_product': by_product, 'by_customer': by_customer
    })


@login_required
def order_ship(request, pk):
    from apps.stock.models import Lot, StockMovement
    from .models import Shipment, ShipmentLine
    from decimal import Decimal

    order = get_object_or_404(SalesOrder, pk=pk)
    
    if order.status in ['shipped', 'cancelled']:
        messages.warning(request, "Cette commande n'est plus expédiable.")
        return redirect('sales:order_detail', pk=order.pk)

    # Préparer les lots disponibles pour chaque produit de la commande
    material_lots = {}
    for line in order.lines.all():
        if line.shipped_quantity < line.quantity:
            lots = Lot.objects.filter(material=line.material, is_active=True, qc_status='approved', current_quantity__gt=0).order_by('expiry_date')
            material_lots[line.material.pk] = lots

    if request.method == 'POST':
        data = request.POST
        
        # 1. Créer l'entête d'expédition
        shipment = Shipment.objects.create(
            order=order,
            shipment_date=data.get('shipment_date'),
            tracking_number=data.get('tracking_number', ''),
            notes=data.get('notes', ''),
            shipped_by=request.user
        )
        
        all_lines_shipped = True
        
        for line in order.lines.all():
            if line.shipped_quantity >= line.quantity:
                continue
                
            qty_str = data.get(f'qty_{line.pk}')
            lot_id = data.get(f'lot_{line.pk}')
            
            if qty_str and lot_id:
                qty = Decimal(qty_str)
                if qty <= 0:
                    continue
                    
                lot = get_object_or_404(Lot, pk=lot_id)
                
                # Vérifier si le stock est suffisant ET approuvé
                if lot.qc_status != 'approved':
                    messages.error(request, f"Le lot {lot.lot_number} est en statut '{lot.get_qc_status_display()}' et ne peut pas être expédié !")
                    shipment.delete()
                    return redirect('sales:order_ship', pk=order.pk)
                    
                if lot.current_quantity < qty:
                    messages.error(request, f"Stock insuffisant pour le lot {lot.lot_number} ({lot.current_quantity} disponibles).")
                    shipment.delete() # Annuler l'expédition entière
                    return redirect('sales:order_ship', pk=order.pk)
                
                # Déduction du stock
                lot.current_quantity -= qty
                lot.save()
                
                # Tracer le mouvement
                movement = StockMovement.objects.create(
                    lot=lot,
                    movement_type='sale',
                    quantity=-qty,
                    reference=f"EXP-{shipment.id} / CV-{order.order_number}",
                    performed_by=request.user
                )
                
                # Ligne d'expédition
                ShipmentLine.objects.create(
                    shipment=shipment,
                    order_line=line,
                    lot=lot,
                    quantity=qty,
                    created_lot_movement=movement
                )
                
                line.shipped_quantity += qty
                line.save()
                
            if line.shipped_quantity < line.quantity:
                all_lines_shipped = False

        if all_lines_shipped:
            order.status = 'shipped'
        elif order.status == 'draft':
            order.status = 'confirmed'
            
        order.save()
        messages.success(request, f"Expédition enregistrée et stock déduit avec succès.")
        return redirect('sales:order_detail', pk=order.pk)

    return render(request, 'sales/order_ship.html', {
        'order': order,
        'material_lots': material_lots
    })


# ─── Export Excel ─────────────────────────────────────────────────────────────

@login_required
def sales_report_excel(request):
    """
    GET /sales/reports/excel/
    Exporte toutes les commandes ventes en fichier Excel professionnel.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        from django.http import HttpResponse
        return HttpResponse(
            "❌ Package 'openpyxl' non installé. Exécutez : pip install openpyxl",
            status=500
        )

    from django.http import HttpResponse
    from django.utils import timezone

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport Ventes CosmoERP"

    # ── Styles ────────────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11, name='Calibri')
    header_fill = PatternFill("solid", fgColor="1B4F72")
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    data_font = Font(size=10, name='Calibri')
    alt_fill = PatternFill("solid", fgColor="EBF5FB")

    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # ── Titre de rapport ──────────────────────────────────────────────────────
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"CosmoERP — Rapport Ventes — Exporté le {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    title_cell.font = Font(bold=True, size=14, color="1B4F72", name='Calibri')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # ── En-têtes colonnes ─────────────────────────────────────────────────────
    headers = ['Date', 'N° Commande', 'Client', 'Produit', 'Qté commandée',
               'Qté expédiée', 'Prix unitaire (DT)', 'Total HT (DT)', 'Statut']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[2].height = 25

    # ── Données ───────────────────────────────────────────────────────────────
    orders = SalesOrder.objects.select_related('customer').prefetch_related('lines__material').all()
    row_idx = 3
    for order in orders:
        for line_idx, line in enumerate(order.lines.all()):
            fill = alt_fill if row_idx % 2 == 0 else None
            row_data = [
                order.order_date.strftime('%d/%m/%Y') if order.order_date else '',
                f"CV-{order.order_number}",
                order.customer.name,
                line.material.name if line.material else '—',
                float(line.quantity),
                float(line.shipped_quantity),
                float(line.unit_price),
                float(line.total_price),
                order.get_status_display(),
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = data_font
                cell.border = border
                if fill:
                    cell.fill = fill
                # Format nombres
                if col in [5, 6, 7, 8]:
                    cell.number_format = '#,##0.00'
            row_idx += 1

    # ── Totaux ────────────────────────────────────────────────────────────────
    if row_idx > 3:
        total_row = row_idx
        ws.cell(row=total_row, column=7, value='TOTAL').font = Font(bold=True, color="1B4F72", name='Calibri')
        total_formula = ws.cell(
            row=total_row, column=8,
            value=f"=SUM(H3:H{total_row - 1})"
        )
        total_formula.font = Font(bold=True, color="1B4F72", name='Calibri')
        total_formula.number_format = '#,##0.00 "DT"'

    # ── Largeur colonnes auto ─────────────────────────────────────────────────
    col_widths = [12, 15, 30, 35, 14, 14, 18, 16, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Figer les en-têtes ────────────────────────────────────────────────────
    ws.freeze_panes = 'A3'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = \
        f'attachment; filename="rapport_ventes_cosmo_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


# ─── Export PDF Fiche de Lot ──────────────────────────────────────────────────

@login_required
def batch_pdf(request, pk):
    """
    GET /sales/batches/<pk>/pdf/  (ici depuis sales pour les lots des produits finis)
    Génère une fiche de lot de production en PDF.
    """
    from apps.production.models import ProductionBatch

    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.colors import HexColor, white, black
        from reportlab.lib.units import cm
    except ImportError:
        from django.http import HttpResponse
        return HttpResponse(
            "❌ Package 'reportlab' non installé. Exécutez : pip install reportlab",
            status=500
        )

    from django.http import HttpResponse

    batch = get_object_or_404(ProductionBatch, pk=pk)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = \
        f'attachment; filename="fiche_lot_{batch.batch_number}.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    margin = 2 * cm

    # ── En-tête bleu foncé ────────────────────────────────────────────────────
    header_color = HexColor('#1B4F72')
    p.setFillColor(header_color)
    p.rect(0, height - 85, width, 85, fill=True, stroke=False)

    p.setFillColor(white)
    p.setFont("Helvetica-Bold", 22)
    p.drawString(margin, height - 38, "CosmoERP")
    p.setFont("Helvetica", 13)
    p.drawString(margin, height - 58, "Fiche de Lot de Production")
    p.setFont("Helvetica", 10)
    p.drawRightString(width - margin, height - 38, f"Lot N° {batch.batch_number}")

    # ── Ligne de séparation ────────────────────────────────────────────────────
    p.setStrokeColor(header_color)
    p.setLineWidth(2)
    p.line(margin, height - 100, width - margin, height - 100)

    # ── Informations principales ───────────────────────────────────────────────
    p.setFillColor(black)
    y = height - 130
    line_h = 22

    def draw_field(label, value, y_pos):
        p.setFont("Helvetica-Bold", 10)
        p.setFillColor(header_color)
        p.drawString(margin, y_pos, f"{label} :")
        p.setFont("Helvetica", 10)
        p.setFillColor(black)
        p.drawString(margin + 5 * cm, y_pos, str(value))
        return y_pos - line_h

    y = draw_field("Produit", batch.formulation.product.name if batch.formulation else '—', y)
    y = draw_field("Formulation", batch.formulation.name if batch.formulation else '—', y)
    y = draw_field("Quantité planifiée",
                   f"{batch.planned_quantity} kg", y)
    y = draw_field("Quantité réelle",
                   f"{batch.actual_quantity} kg" if batch.actual_quantity else 'En cours', y)
    y = draw_field("Date de production",
                   batch.production_date.strftime('%d/%m/%Y') if batch.production_date else '—', y)
    y = draw_field("Responsable",
                   batch.responsible.get_full_name() or batch.responsible.username
                   if batch.responsible else '—', y)
    y = draw_field("Statut", batch.get_status_display(), y)
    y -= 10

    # ── Matières premières ────────────────────────────────────────────────────
    p.setFont("Helvetica-Bold", 12)
    p.setFillColor(header_color)
    p.drawString(margin, y, "Matières premières utilisées :")
    y -= 5
    p.setStrokeColor(header_color)
    p.setLineWidth(1)
    p.line(margin, y, width - margin, y)
    y -= line_h

    raw_materials = batch.raw_materials.select_related('material').all()
    if raw_materials:
        # En-têtes tableau
        p.setFont("Helvetica-Bold", 9)
        p.setFillColor(white)
        p.setStrokeColor(header_color)
        p.setFillColor(header_color)
        p.rect(margin, y - 2, width - 2 * margin, 18, fill=True, stroke=False)
        p.setFillColor(white)
        p.drawString(margin + 5, y + 4, "Ingrédient")
        p.drawString(margin + 10 * cm, y + 4, "Requis")
        p.drawString(margin + 14 * cm, y + 4, "Réel utilisé")
        y -= 18 + 5

        for rm in raw_materials:
            name = rm.material.name if rm.material else rm.ingredient_name
            p.setFont("Helvetica", 9)
            p.setFillColor(black)
            p.drawString(margin + 5, y, name[:55])
            p.drawString(margin + 10 * cm, y, f"{rm.required_quantity}")
            p.drawString(margin + 14 * cm, y,
                         f"{rm.actual_quantity}" if rm.actual_quantity else "—")
            y -= line_h - 4
            if y < 5 * cm:
                p.showPage()
                y = height - margin

    # ── Signature ─────────────────────────────────────────────────────────────
    p.setFont("Helvetica", 9)
    p.setFillColor(HexColor('#888888'))
    p.drawString(margin, 2 * cm, f"Généré par CosmoERP le "
                 f"{__import__('django.utils.timezone', fromlist=['timezone']).timezone.now().strftime('%d/%m/%Y %H:%M')}")
    p.drawRightString(width - margin, 2 * cm, "© CosmoERP — Confidentiel")

    p.showPage()
    p.save()
    return response

